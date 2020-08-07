import requests
import time
from openpyxl import Workbook
from openpyxl import load_workbook
arquivo_excel = Workbook()
planilha1 = arquivo_excel.active


caminho = '/files/pythonexcel/myfile.xlsx'
arquivo_excel1 = load_workbook(caminho)

planilha2 = arquivo_excel1.active



c = 2; z = 2; m1 = 1; 

while True:
    c3 = planilha2.cell(column = 1, row = z)
    z +=1


    a_string = (c3.value)
    alphanumeric = ""


    for character in a_string:
        if character.isalnum():
            alphanumeric += character


    r = requests.get('https://www.receitaws.com.br/v1/cnpj/' + (alphanumeric) )
    r_dict = r.json()


    planilha1['A1'] = 'Nome'
    planilha1['b1'] = 'CNPJ'
    planilha1['c1'] = 'CEP'
    planilha1['d1'] = 'Endereço'
    planilha1['e1'] = 'Cidade'
    planilha1['f1'] = 'UF'
    planilha1['g1'] = 'Capital Social'
    planilha1['h1'] = 'Situação'

        
    planilha1.cell(row = (c)  , column = 1, value = r_dict['nome'])

    planilha1.cell(row = (c)  , column = 2, value = r_dict['cnpj'])

    planilha1.cell(row = (c)  , column = 3, value = r_dict['cep'])

    planilha1.cell(row = (c)  , column = 4, value = f"{r_dict['logradouro']}, {r_dict['numero']}, {r_dict['complemento']}")

    planilha1.cell(row = (c)  , column = 5, value = r_dict['municipio'])

    planilha1.cell(row = (c)  , column = 6, value = r_dict['uf'])

    planilha1.cell(row = (c)  , column = 7, value = r_dict['capital_social'])

    planilha1.cell(row = (c)  , column = 8, value = r_dict['situacao'])

    c += 1
    

    arquivo_excel.save("Informações PDG.xlsx") 


    print(f'Empresa {m1}')

    if (m1 % 3 == 0):
        print("Espere um minuto para continuar")
        time.sleep(60)
        m1 += 1
    else:
        m1 += 1